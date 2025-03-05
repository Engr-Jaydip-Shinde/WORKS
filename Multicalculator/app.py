from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os
app = Flask(__name__)

EXCEL_FILE = "form_data.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Responses"
    ws.append(["Name", "Message Type", "Message"])  
    wb.save(EXCEL_FILE)
@app.route('/submit', methods=['POST'])
def submit_form():
    data = request.form
    name = data.get('name')
    message_type = data.get('type')  
    message = data.get('message')
    if not all([name, message_type, message]):
        return jsonify({"status": "error", "message": "All fields are required"}), 400
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, message_type, message])
    wb.save(EXCEL_FILE)
    return jsonify({"status": "success", "message": "Form data saved successfully!"})
if __name__ == '__main__':
    app.run(debug=True)  
